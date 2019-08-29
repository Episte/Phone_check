import openpyxl
import checkm_1
import csv
#結果記入ファイルを読み込み
result_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\result1.xlsx"
result_wb = openpyxl.load_workbook(result_file)
#シートの読み込み
working_phone_ws = result_wb["Working_Phone"]
no_dhcp_ws = result_wb["No_DHCP"]
no_cucm_ws = result_wb["No_CUCM"]
#lease情報のexcelファイル読み込み
lease_file = open(r"C:\Users\石川武\Desktop\pyproject\phonecheck\lease.csv","r",encoding="utf_8",errors="",newline="")
lease_data = csv.reader(lease_file)
#phone情報のexcelファイル読み込み
phone_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\cucm_phone.xlsx"
phone_wb = openpyxl.load_workbook(phone_file)
phone_ws = phone_wb["Sheet1"]
#リース状態&ATA以外の辞書作成
lease_dic = {}
for i in lease_data:
    if (i[3] == "LEASE") and (not i[6].startswith("ATA")):
        lease_dic[i[6]] = i[0]
#リース状態&ATA以外のホスト名を取得
lease_hosts = [i for i in lease_dic]
#CUCMの辞書作成
cucm_dictionary = {}
for i in range(2,phone_ws.max_row+1):
    cucm_dictionary[phone_ws.cell(row=i,column=3).value.rstrip("\n")]\
     = phone_ws.cell(row=i,column=8).value.rstrip("\n")
#cucmのホスト名を取得
cucm_hosts = [i for i in cucm_dictionary]
#登録機器の差分を計算、表示する
#DHCPなし、CUCMあり
data_onlys = list(set(cucm_hosts)-set(lease_hosts))
#DHCP、CUCMあり
match_phones = list(set(cucm_hosts) and set(lease_hosts))
#DHCPあり、CUCMなし
lease_onlys = list(set(lease_hosts)-set(cucm_hosts))
#DHCPあり、CUCMありのIP電話の正常性確認
print("\n\nDHCPからアドレスが割り当てられたIP電話の正常性確認\n\n")
working_phone_row = 2
for match_phone in match_phones:
    working_phone_ws.cell(row=working_phone_row,column=1).value = match_phone
    #パターン1
    if cucm_dictionary[match_phone] != "なし" and cucm_dictionary[match_phone] != "不明":
        working_phone_ws.cell(row=working_phone_row,column=2).value = cucm_dictionary[match_phone]
        #疎通がとれるかチェック
        if checkm_1.Ping(cucm_dictionary[match_phone]):
            result = checkm_1.Phone_Check(cucm_dictionary[match_phone])
            working_phone_ws.cell(row=working_phone_row,column=3).value = "○"
            for i in range(len(result)):
                working_phone_ws.cell(row=working_phone_row,column=i+4).value\
                = result[i]
        #疎通取れなかった
        else:
            for i in range(3,10):
                working_phone_ws.cell(row=working_phone_row,column=i).value = "×"
    #パターン2
    else:
        working_phone_ws.cell(row=working_phone_row,column=2).value = lease_dic[match_phone]
        #疎通がとれるかチェック
        if checkm_1.Ping(lease_dic[match_phone]):
            result = checkm_1.Phone_Check(lease_dic[match_phone])
            working_phone_ws.cell(row=working_phone_row,column=3).value = "○"
            for i in range(len(result)):
                working_phone_ws.cell(row=working_phone_row,column=i+4).value\
                = result[i]
        #疎通取れなかった
        else:
            for i in range(3,10):
                working_phone_ws.cell(row=working_phone_row,column=i).value = "×"
    working_phone_row += 1
#DHCPなし、CUCMありのIP電話の正常性確認
print("-------------------------------------")
print("DHCPのリース情報に載っていない機器の正常性チェック")
no_dhcp_row = 2
for data_only in data_onlys:
    if cucm_dictionary[data_only] == "なし" or cucm_dictionary[data_only] == "不明":
        print(data_only + "はデータはCUCMに存在しますが、IPアドレスがありません")
        no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only
        for i in range(2,10):
            no_dhcp_ws.cell(row=no_dhcp_row,column=i).value = "×"
    else:
        if checkm_1.Ping(cucm_dictionary[data_only]):
            no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only
            result_dhcp = checkm_1.Phone_Check(cucm_dictionary[data_only])
            for i in range(len(result_dhcp)):
                no_dhcp_ws.cell(row=no_dhcp_row,column=i+4).value = result_dhcp[i]
        else:
            no_dhcp_ws.cell(row=no_dhcp_row,column=2).value = cucm_dictionary[data_only]
            for i in range(3,10):
                no_dhcp_ws.cell(row=no_dhcp_row,column=i).value = "×"
    no_dhcp_row += 1

#DHCPあり、CUCMなしのIP電話の正常性確認
print("\n\n-------------------------------------")
print("以下はDHCPからIPを割り当てられているが、CUCMには登録されていない機器です")
no_cucm_row = 2
for lease_only in lease_onlys:
    if cucm_dictionary[lease_only] != "なし" and cucm_dictionary[lease_only] != "不明":
        no_cucm_ws.cell(row=no_cucm_row,column=1).value = lease_only
        no_cucm_ws.cell(row=no_cucm_row,column=2).value = lease_dic[lease_only]
        no_cucm_row += 1
#保存して終了
result_wb.save(result_file)
