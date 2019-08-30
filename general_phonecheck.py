import os,sys
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import checkm_1
import csv

#button1のコマンド:csvファイル名をゲット
def csv_clicked():
    ftype1 = [("csvファイル","*.csv")]
    dir1 = 'C:'
    file1 = filedialog.askopenfilename(filetypes=ftype1,initialdir=dir1)
    entrybox1.insert(tk.END,file1)
#button2のコマンド:excelファイル名をゲット
def excel_clicked():
    ftype2 = [("Excelファイル","*.xlsx")]
    dir2 = 'C:'
    file2 = filedialog.askopenfilename(filetypes=ftype2,initialdir=dir2)
    entrybox2.insert(tk.END,file2)
#button3のコマンド:結果ファイル用のディレクトリゲット
def dir_clicked():
    dir3 = 'C:'
    file3 = filedialog.askdirectory(initialdir=dir3)
    entrybox3.insert(tk.END,file3)

def process():
    #結果記入ファイルを作成
    result_wb = openpyxl.Workbook()
    #シート名変更
    working_phone_ws = result_wb.active
    working_phone_ws.title = "Working_Phone"
    #シート作成
    no_dhcp_ws = result_wb.create_sheet(title="No_DHCP")
    no_cucm_ws = result_wb.create_sheet(title="No_CUCM")
    #各シートに項目を作成
    check_items = ["ホスト名","IPアドレス","登録","DHCPサーバー","TFTPサーバー1",\
                   "TFTPサーバー2","CUCMサーバー1","CUCMサーバー2","ITLファイル"]
    for i in range(9):
        working_phone_ws.cell(row=1,column=i+1).value = check_items[i]
        no_dhcp_ws.cell(row=1,column=i+1).value = check_items[i]
    for i in range(2):
        no_cucm_ws.cell(row=1,column=i+1).value = check_items[i]
    #lease情報のexcelファイル読み込み
    lease_file = open(entrybox1.get(),"r",encoding="utf_8",errors="",newline="")
    lease_data = csv.reader(lease_file)
    #phone情報のexcelファイル読み込み
    phone_file = entrybox2.get()
    phone_wb = openpyxl.load_workbook(phone_file)
    phone_ws = phone_wb["Sheet1"]
    #リース状態&ATA以外の辞書作成
    lease_dic = {}
    for i in lease_data:
        if (i[3] == "LEASE") and (not i[6].startswith("ATA")):
            lease_dic[i[6]] = i[0]
    #リース状態&ATA以外のホスト名を取得
    lease_hosts = [i for i in lease_dic]
    #CUCMの辞書作成
    cucm_dictionary = {}
    for i in range(1,phone_ws.max_row+1):
        cucm_dictionary[phone_ws.cell(row=i,column=3).value.rstrip("\n")]\
         = phone_ws.cell(row=i,column=8).value.rstrip("\n")
    #cucmのホスト名を取得
    cucm_hosts = [i for i in cucm_dictionary]
    #登録機器の差分を計算、表示する
    #DHCPなし、CUCMあり
    data_onlys = list(set(cucm_hosts)-set(lease_hosts))
    #DHCP、CUCMあり
    match_phones = list(set(cucm_hosts) and set(lease_hosts))
    #DHCPあり、CUCMなし
    lease_onlys = list(set(lease_hosts)-set(cucm_hosts))
    #DHCPあり、CUCMありのIP電話の正常性確認
    print("\n\nDHCPからアドレスが割り当てられたIP電話の正常性確認\n\n")
    working_phone_row = 2
    for match_phone in match_phones:
        working_phone_ws.cell(row=working_phone_row,column=1).value = match_phone
        #パターン1
        if cucm_dictionary[match_phone] != "なし" and cucm_dictionary[match_phone] != "不明":
            working_phone_ws.cell(row=working_phone_row,column=2).value = cucm_dictionary[match_phone]
            #疎通がとれるかチェック
            if checkm_1.Ping(cucm_dictionary[match_phone]):
                result = checkm_1.Phone_Check(cucm_dictionary[match_phone])
                working_phone_ws.cell(row=working_phone_row,column=3).value = "○"
                for i in range(len(result)):
                    working_phone_ws.cell(row=working_phone_row,column=i+4).value\
                    = result[i]
            #疎通取れなかった
            else:
                for i in range(3,10):
                    working_phone_ws.cell(row=working_phone_row,column=i).value = "×"
        #パターン2
        else:
            working_phone_ws.cell(row=working_phone_row,column=2).value = lease_dic[match_phone]
            #疎通がとれるかチェック
            if checkm_1.Ping(lease_dic[match_phone]):
                result = checkm_1.Phone_Check(lease_dic[match_phone])
                working_phone_ws.cell(row=working_phone_row,column=3).value = "○"
                for i in range(len(result)):
                    working_phone_ws.cell(row=working_phone_row,column=i+4).value\
                    = result[i]
            #疎通取れなかった
            else:
                for i in range(3,10):
                    working_phone_ws.cell(row=working_phone_row,column=i).value = "×"
        working_phone_row += 1
    #DHCPなし、CUCMありのIP電話の正常性確認
    print("-------------------------------------")
    print("DHCPのリース情報に載っていない機器の正常性チェック")
    no_dhcp_row = 2
    for data_only in data_onlys:
        if cucm_dictionary[data_only] == "なし" or cucm_dictionary[data_only] == "不明":
            print(data_only + "はデータはCUCMに存在しますが、IPアドレスがありません")
            no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only
            for i in range(2,10):
                no_dhcp_ws.cell(row=no_dhcp_row,column=i).value = "×"
        else:
            if checkm_1.Ping(cucm_dictionary[data_only]):
                no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only
                result_dhcp = checkm_1.Phone_Check(cucm_dictionary[data_only])
                for i in range(len(result_dhcp)):
                    no_dhcp_ws.cell(row=no_dhcp_row,column=i+4).value = result_dhcp[i]
            else:
                no_dhcp_ws.cell(row=no_dhcp_row,column=2).value = cucm_dictionary[data_only]
                for i in range(3,10):
                    no_dhcp_ws.cell(row=no_dhcp_row,column=i).value = "×"
        no_dhcp_row += 1

    #DHCPあり、CUCMなしのIP電話の正常性確認
    print("\n\n-------------------------------------")
    print("以下はDHCPからIPを割り当てられているが、CUCMには登録されていない機器です")
    no_cucm_row = 2
    for lease_only in lease_onlys:
        if cucm_dictionary[lease_only] != "なし" and cucm_dictionary[lease_only] != "不明":
            no_cucm_ws.cell(row=no_cucm_row,column=1).value = lease_only
            no_cucm_ws.cell(row=no_cucm_row,column=2).value = lease_dic[lease_only]
            no_cucm_row += 1
    #保存して終了
    result_filename = "/" + entrybox4.get()
    result_filename.rstrip(".xlsx")
    result_ex = ".xlsx"
    result_count = 0
    filepath = entrybox3.get() + result_filename + result_ex
    while os.path.exists(filepath):
        int(result_count)
        result_count += 1
        filepath = entrybox3.get() + result_filename + str(result_count) + result_ex
    result_wb.save(filepath)
    messagebox.showinfo("確認","処理が終了しました。ファイルを確認してください。")
#サイズ要調整
root = tk.Tk()
root.title("IP電話チェックツール")
root.geometry("450x200")
root.resizable(0,0)
#frame1作成
frame1 = tk.Frame(root)
#1行目の部品作成
file1 = ""
label1 = tk.Label(frame1,text="csvファイル")
entrybox1 = tk.Entry(frame1,width=50)
button1 = tk.Button(frame1,text="参照",command=csv_clicked)
#frame1と1行目を表示
frame1.grid()
label1.grid(row=0,column=0,padx=5,pady=5)
entrybox1.grid(row=0,column=1,padx=5,pady=5)
button1.grid(row=0,column=2,padx=5,pady=5)
#2行目の部品
label2 = tk.Label(frame1,text="Excelファイル")
entrybox2 = tk.Entry(frame1,width=50)
button2 = tk.Button(frame1,text="参照",command=excel_clicked)
#2行目を表示
label2.grid(row=1,column=0,padx=5,pady=5)
entrybox2.grid(row=1,column=1,padx=5,pady=5)
button2.grid(row=1,column=2,padx=5,pady=5)
#3行目の部品
label3 = tk.Label(frame1,text="結果ファイル")
entrybox3 = tk.Entry(frame1,width=50)
button3 = tk.Button(frame1,text="参照",command=dir_clicked)
#3行目を表示
label3.grid(row=2,column=0,padx=5,pady=5)
entrybox3.grid(row=2,column=1,padx=5,pady=5)
button3.grid(row=2,column=2,padx=5,pady=5)
#4行目の部品
label4 = tk.Label(frame1,text="ファイル名")
entrybox4 = tk.Entry(frame1,width=50)
#4行目の表示
label4.grid(row=3,column=0,padx=5,pady=5)
entrybox4.grid(row=3,column=1,padx=5,pady=5)
#処理開始ボタン作成
frame2 = tk.Frame(root)
start_button = tk.Button(frame2,text="処理開始",command=process)
#処理開始ボタン表示
frame2.grid()
start_button.grid(row=2,column=1,padx=5,pady=5)

root.mainloop()
