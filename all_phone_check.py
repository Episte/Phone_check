import re
import pings
import urllib.request
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd

#引数に指定したipアドレスにpingし、疎通可能性を戻り値として返す関数
def Ping(ip):
    p = pings.Ping()
    res = p.ping(ip,times=4)
    print("-------------------------------------")
    res.print_messages()
    print("-------------------------------------")
    return res.is_reached()

def Phone_Check(ip):
    #webスクレイピングでデータ取得
    weburl = r"http://"+ip+"/CGI/Java/Serviceability?adapter=device.statistics.configuration"
    with urllib.request.urlopen(weburl) as source:
        html = source.read()
    #データを処理しやすく加工
    soup = BeautifulSoup(html,"lxml")
    settings = soup.find_all("b")
    txtdata = ""
    for data in settings:
        txtdata += " " + data.text.strip()
    #DHCPサーバを判定
    dhcp = re.search("DHCPサーバ 192.168.241.12",txtdata)
    if dhcp:
        print(dhcp.group())
        dhcpjudge = "OK"
    else:
        print("DHCPに問題があります")
        dhcpjudge = "NO"
    #TFTPサーバ1を判定
    tftp1 = re.search("TFTPサーバ1 192.168.241.35",txtdata)
    if tftp1:
        print(tftp1.group())
        tftp1judge = "OK"
    else:
        print("TFTP1に問題があります")
        tftp1judge = "NO"
    #CUCMサーバ1を判定
    cucm1 = re.search("CUCMサーバ1 192.168.241.34  Active",txtdata)
    #CUCMサーバ2を判定
    cucm2 = re.search("CUCMサーバ2 192.168.241.35  Standby",txtdata)
    if cucm1:
        print(cucm1.group())
        cucm1judge = "OK"
    else:
        cucm1judge = "NO"
    if cucm2:
        print(cucm2.group())
        cucm2judge = "OK"
    else:
        cucm2judge = "NO"
    if not (cucm1 and cucm2):
        print("正しいCUCMに登録されていません")
    #TVS判定
    tvs = re.search("TVS",txtdata)
    if tvs:
        print("ITLファイルインストール済")
        tvsjudge = "ITLファイルインストール済"
    else:
        print("ITLファイルがインストールされていません")
        tvsjudge = "ITLファイル未インストール"
    if dhcp and tftp1 and cucm1 and cucm2 and tvs:
        print("機器" + ip + "は正常に作動しています\n\n")
    else:
        print("機器" + ip + "に異常があります\n\n")
    return [dhcpjudge,tftp1judge,cucm1judge,cucm2judge,tvsjudge]
#結果ファイルを用意
result_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\result1.xlsx"
result_wb = openpyxl.load_workbook(result_file)
#シートの読み込み
working_phone_ws = result_wb["Working_Phone"]
no_dhcp_ws = result_wb["No_DHCP"]
no_cucm_ws = result_wb["No_CUCM"]
#lease情報のexcelファイル読み込み
lease_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\leasedata.xlsx"
lease_wb = openpyxl.load_workbook(lease_file)
lease_ws = lease_wb["Sheet1"]
#phone情報のexcelファイル読み込み
phone_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\cucm_phone.xlsx"
phone_wb = openpyxl.load_workbook(phone_file)
phone_ws = phone_wb["Sheet1"]
#リース状態のipアドレスを取得
ips = [lease_ws.cell(row=i,column=1).value for i in range(2,lease_ws.max_row+1) \
if lease_ws.cell(row=i,column=4).value == "LEASE"]
#リース状態のホスト名を取得
lease_hosts = [lease_ws.cell(row=i,column=7).value for i in range(2,lease_ws.max_row+1) \
if lease_ws.cell(row=i,column=4).value == "LEASE"]
#CUCMの電話データのホスト名とipアドレスの辞書取作成
phone_dictionary = {}
for i in range(2,phone_ws.max_row+1):
    phone_dictionary[phone_ws.cell(row=i,column=3).value.rstrip("\n")]\
     = phone_ws.cell(row=i,column=8).value.rstrip("\n")
#cucmのホスト名を取得
phone_hosts = [i for i in phone_dictionary]
#ホスト名を昇順にソートする
lease_hosts.sort()
phone_hosts.sort()
#登録機器の差分を計算、表示する
#DHCPからアドレスを払い出されていないが、データのみ存在するIP電話のリスト
data_only_phones = list(set(phone_hosts)-set(lease_hosts))
#登録されていないIP電話のリスト
unregist_phones = list(set(lease_hosts)-set(phone_hosts))
#DHCPから割り当てられており、登録もされているIP電話のリスト
match_phones = list(set(phone_hosts) and set(lease_hosts))
#取得したipを関数に渡す
print("\n\nDHCPからアドレスが割り当てられたIP電話の正常性確認\n\n")
working_phone_row = 2
for match_phone in match_phones:
    if phone_dictionary[match_phone] != "なし":
        if Ping(phone_dictionary[match_phone]):
            result = Phone_Check(phone_dictionary[match_phone])
            working_phone_ws.cell(row=working_phone_row,column=1).value = match_phone
            working_phone_ws.cell(row=working_phone_row,column=2).value = phone_dictionary[match_phone]
            for i in range(5):
                working_phone_ws.cell(row=working_phone_row,column=i+3).value\
                = result[i]
            working_phone_row += 1
#DHCPからアドレスが割り当てられていないがcucmに登録されている電話の正常性
print("-------------------------------------")
print("DHCPのリース情報に載っていない機器の正常性チェック")
no_dhcp_row = 2
for data_only_phone in data_only_phones:
    if phone_dictionary[data_only_phone] == "なし":
        print(data_only_phone + "はデータはCUCMに存在しますが、IPアドレスがありません")
        no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only_phone
        no_dhcp_row += 1
    elif Ping(phone_dictionary[data_only_phone]):
        no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only_phone
        result_dhcp = Phone_Check(phone_dictionary[data_only_phone])
        for i in range(5):
            no_dhcp_ws.cell(row=no_dhcp_row,column=i+2).value = result_dhcp[i]
        no_dhcp_row += 1
#DHCPにデータはあるがCUCMに登録されていない機器
print("\n\n-------------------------------------")
print("以下はDHCPからIPを割り当てられているが、CUCMには登録されていない機器です")
unregists = [lease_host for lease_host in lease_hosts if phone_dictionary[lease_host]=="なし"]
print(unregists)
no_cucm_row = 2
for unregist in unregists:
    no_cucm_ws.cell(row=no_cucm_row,column=1).value = unregist
    no_cucm_row += 1
#保存して終了
result_wb.save(result_file)
