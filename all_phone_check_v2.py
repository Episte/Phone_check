import openpyxl
import pandas as pd
import check

#結果記入ファイルを用意
result_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\result1.xlsx"
result_wb = openpyxl.load_workbook(result_file)
#シートの読み込み
working_phone_ws = result_wb["Working_Phone"]
no_dhcp_ws = result_wb["No_DHCP"]
no_cucm_ws = result_wb["No_CUCM"]
#lease情報のexcelファイル読み込み
lease_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\leasedata.xlsx"
lease_wb = openpyxl.load_workbook(lease_file)
lease_ws = lease_wb["Sheet1"]
#phone情報のexcelファイル読み込み
phone_file = r"C:\Users\石川武\Desktop\pyproject\phonecheck\cucm_phone.xlsx"
phone_wb = openpyxl.load_workbook(phone_file)
phone_ws = phone_wb["Sheet1"]
#リース状態のipアドレスを取得
ips = [lease_ws.cell(row=i,column=1).value for i in range(2,lease_ws.max_row+1) \
if lease_ws.cell(row=i,column=4).value == "LEASE"]
#リース状態のホスト名を取得
lease_hosts = [lease_ws.cell(row=i,column=7).value for i in range(2,lease_ws.max_row+1) \
if lease_ws.cell(row=i,column=4).value == "LEASE"]
#CUCMの電話データのホスト名とipアドレスの辞書取作成
phone_dictionary = {}
for i in range(2,phone_ws.max_row+1):
    phone_dictionary[phone_ws.cell(row=i,column=3).value.rstrip("\n")]\
     = phone_ws.cell(row=i,column=8).value.rstrip("\n")
#cucmのホスト名を取得
phone_hosts = [i for i in phone_dictionary]
#ホスト名を昇順にソートする
lease_hosts.sort()
phone_hosts.sort()
#登録機器の差分を計算、表示する
#CUCMあり、DHCPなし
data_only_phones = list(set(phone_hosts)-set(lease_hosts))
#DHCPあり、CUCMなし
unregist_phones = list(set(lease_hosts)-set(phone_hosts))
#DHCP、CUCMあり
match_phones = list(set(phone_hosts) and set(lease_hosts))
#DHCPあり、CUCMありのIP電話の正常性確認
print("\n\nDHCPからアドレスが割り当てられたIP電話の正常性確認\n\n")
working_phone_row = 2
for match_phone in match_phones:
    if phone_dictionary[match_phone] != "なし":
        if check.Ping(phone_dictionary[match_phone]):
            result = check.Phone_Check(phone_dictionary[match_phone])
            working_phone_ws.cell(row=working_phone_row,column=1).value = match_phone
            working_phone_ws.cell(row=working_phone_row,column=2).value = phone_dictionary[match_phone]
            for i in range(5):
                working_phone_ws.cell(row=working_phone_row,column=i+3).value\
                = result[i]
            working_phone_row += 1
#DHCPなし、CUCMありのIP電話の正常性確認
print("-------------------------------------")
print("DHCPのリース情報に載っていない機器の正常性チェック")
no_dhcp_row = 2
for data_only_phone in data_only_phones:
    if phone_dictionary[data_only_phone] == "なし":
        print(data_only_phone + "はデータはCUCMに存在しますが、IPアドレスがありません")
        no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only_phone
        no_dhcp_row += 1
    elif check.Ping(phone_dictionary[data_only_phone]):
        no_dhcp_ws.cell(row=no_dhcp_row,column=1).value = data_only_phone
        result_dhcp = check.Phone_Check(phone_dictionary[data_only_phone])
        for i in range(5):
            no_dhcp_ws.cell(row=no_dhcp_row,column=i+2).value = result_dhcp[i]
        no_dhcp_row += 1
#DHCPあり、CUCMなしのIP電話の正常性確認
print("\n\n-------------------------------------")
print("以下はDHCPからIPを割り当てられているが、CUCMには登録されていない機器です")
unregists = [lease_host for lease_host in lease_hosts if phone_dictionary[lease_host]=="なし"]
print(unregists)
no_cucm_row = 2
for unregist in unregists:
    no_cucm_ws.cell(row=no_cucm_row,column=1).value = unregist
    no_cucm_row += 1
#保存して終了
result_wb.save(result_file)
